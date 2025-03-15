from thefuzz import fuzz
from thefuzz import process
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import PayerGroups, Payers, PayerDetails
from django.shortcuts import render
import logging
import openpyxl
import re

logger = logging.getLogger(__name__)

def upload_page(request):
    return render(request, 'upload.html')

from django.shortcuts import render
from .models import PayerGroups, Payers, PayerDetails

def view_mappings(request):
    payer_details = PayerDetails.objects.all()
    payers = Payers.objects.all()
    payer_groups = PayerGroups.objects.all()

    context = {
        'payer_details': payer_details,
        'payers': payers,
        'payer_groups': payer_groups,
    }
    return render(request, 'view_mappings.html', context)

def view_payer_details(request):
    payer_details = PayerDetails.objects.all()
    context = {
        'payer_details': payer_details,
    }
    return render(request, 'view_payer_details.html', context)

def view_payers(request):
    payers = Payers.objects.all()
    context = {
        'payers': payers,
    }
    return render(request, 'view_payers.html', context)

def view_payer_groups(request):
    payer_groups = PayerGroups.objects.all()
    context = {
        'payer_groups': payer_groups,
    }
    return render(request, 'view_payer_groups.html', context)

def view_payer_mappings(request):
    payers = Payers.objects.all()
    payer_mappings = []
    for payer in payers:
        payer_mappings.append({
            'payer_name': payer.name,
            'duplicate_names': [detail.name for detail in payer.details.all()],
        })
    context = {
        'payer_mappings': payer_mappings,
    }
    return render(request, 'view_payer_mappings.html', context)

def view_payer_group_mappings(request):
    payer_groups = PayerGroups.objects.all()
    payer_group_mappings = []
    for group in payer_groups:
        payer_group_mappings.append({
            'payer_group_name': group.name,
            'payers': [payer.name for payer in group.payers.all()],
        })
    context = {
        'payer_group_mappings': payer_group_mappings,
    }
    return render(request, 'view_payer_group_mappings.html', context)

class UploadView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            PayerDetails.objects.all().delete()
            Payers.objects.all().delete()
            PayerGroups.objects.all().delete()

            uploaded_file = request.FILES.get('file')
            if not uploaded_file:
                return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

            file_name = uploaded_file.name
            if not file_name.endswith('.xlsx'):
                return Response({"error": "Unsupported file format. Please upload an XLSX file."}, status=status.HTTP_400_BAD_REQUEST)

            self.process_xlsx(uploaded_file)

            return Response({"message": "File processed successfully."}, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error processing file: {e}", exc_info=True)
            return Response({"error": "An error occurred while processing the file."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def process_xlsx(self, file):
        try:
            workbook = openpyxl.load_workbook(file)
            excluded_sheets = ['Legend', 'Legend (1)', 'OpenDental']

            for sheet_name in workbook.sheetnames:
                if sheet_name in excluded_sheets:
                    logger.info(f"Skipping excluded sheet: {sheet_name}")
                    continue

                sheet = workbook[sheet_name]
                rows = sheet.iter_rows(values_only=True)

                try:
                    headers = next(rows)
                    logger.debug(f"Headers in sheet '{sheet_name}': {headers}")
                except StopIteration:
                    logger.warning(f"Skipping sheet '{sheet_name}' because it is empty.")
                    continue

                payer_name_col = self.find_column_index(headers, ['Payer Name', 'Name', 'Payer', 'Payer Identification Information'])
                payer_id_col = self.find_column_index(headers, ['Payer ID', 'ID'])

                if payer_name_col is None or payer_id_col is None:
                    logger.warning(f"Skipping sheet '{sheet_name}' because required columns are missing.")
                    continue

                for row_index, row in enumerate(rows, start=2):
                    try:
                        payer_name = self.clean_value(row[payer_name_col])
                        payer_id = self.clean_value(row[payer_id_col])

                        if not payer_name or not payer_id:
                            logger.warning(f"Skipping row {row_index} in sheet '{sheet_name}' because of missing data.")
                            continue

                        logger.debug(f"Processing row {row_index} in sheet '{sheet_name}': Payer Name = {payer_name}, Payer ID = {payer_id}")

                        self.process_payer(payer_name, payer_id)
                    except Exception as e:
                        logger.error(f"Error processing row {row_index} in sheet '{sheet_name}': {e}")

        except Exception as e:
            logger.error(f"Error processing XLSX file: {e}", exc_info=True)
            raise

    def find_column_index(self, headers, possible_names):
        for name in possible_names:
            if name in headers:
                return headers.index(name)
        return None

    def clean_value(self, value):
        if value and isinstance(value, str):
            value = re.sub(r'^="([^"]*)"$', r'\1', value)
            value = value.strip()
        return value

    def process_payer(self, payer_name, payer_id):
        logger.debug(f"Processing payer: Name = {payer_name}, ID = {payer_id}")

        payer_group = self.find_or_create_payer_group(payer_name)

        payer = self.find_or_create_payer(payer_name, payer_group)

        self.create_payer_details(payer, payer_name, payer_id)

    def find_or_create_payer_group(self, payer_name):
        payer_group_name = self.extract_payer_group_name(payer_name)
        logger.debug(f"Extracted payer group name: {payer_group_name}")

        payer_groups = PayerGroups.objects.all()
        payer_group_names = [group.name for group in payer_groups]

        if payer_group_names:
            best_match, score = process.extractOne(payer_group_name, payer_group_names, scorer=fuzz.token_sort_ratio)
        else:
            best_match, score = None, 0

        if score >= 80:
            payer_group = payer_groups.get(name=best_match)
            logger.debug(f"Found similar payer group: {payer_group}")
        else:
            payer_group = PayerGroups.objects.create(name=payer_group_name)
            logger.debug(f"Created new payer group: {payer_group}")

        return payer_group

    def find_or_create_payer(self, payer_name, payer_group):
        payers = Payers.objects.filter(payer_group=payer_group)
        payer_names = [payer.name for payer in payers]

        if payer_names:
            best_match, score = process.extractOne(payer_name, payer_names, scorer=fuzz.token_sort_ratio)
        else:
            best_match, score = None, 0

        if score >= 80:
            payer = payers.get(name=best_match)
            logger.debug(f"Found similar payer: {payer}")
        else:
            payer = Payers.objects.create(name=payer_name, payer_group=payer_group)
            logger.debug(f"Created new payer: {payer}")

        return payer

    def create_payer_details(self, payer, payer_name, payer_id):
        payer_detail, created = PayerDetails.objects.get_or_create(
            payer=payer,
            name=payer_name,
            payer_number=payer_id,
            defaults={'tax_id': None}
        )
        logger.debug(f"Payer Detail: {payer_detail}, Created: {created}")

    def extract_payer_group_name(self, payer_name):
        return payer_name.split()[0]

